# controllers/export_controller.py
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from tkinter import messagebox

def export_treeview_to_excel(treeview, filename):
    try:
        rows = [treeview.item(item)['values'] for item in treeview.get_children()]
        if not rows:
            messagebox.showwarning("Exportar Excel", "No hay datos para exportar.")
            return

        columns = treeview['columns']

        wb = Workbook()
        ws = wb.active
        ws.title = "Datos"

        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=column_title)
            cell.font = cell.font.copy(bold=True)

        for row_num, row_data in enumerate(rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        for col_num, column_title in enumerate(columns, 1):
            max_length = len(str(column_title))
            for row_num in range(2, len(rows) + 2):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        wb.save(filename)
        messagebox.showinfo("Exportar Excel", f"Datos exportados exitosamente a:\n{filename}")
    except Exception as e:
        messagebox.showerror("Exportar Excel", f"Error al exportar: {e}")

def export_treeview_to_pdf(treeview, filename, title="Reporte"):
    try:
        rows = [treeview.item(item)['values'] for item in treeview.get_children()]
        if not rows:
            messagebox.showwarning("Exportar PDF", "No hay datos para exportar.")
            return

        columns = treeview['columns']

        doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 12))

        data = [list(columns)] + rows

        table = Table(data)
        style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ])
        table.setStyle(style)

        elements.append(table)
        doc.build(elements)

        messagebox.showinfo("Exportar PDF", f"Datos exportados exitosamente a:\n{filename}")
    except Exception as e:
        messagebox.showerror("Exportar PDF", f"Error al exportar: {e}")